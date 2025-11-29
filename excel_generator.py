"""Excel Report Generation Module for PMO Reports"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl


def create_excel_report(projects, output_path):
    """Generate a comprehensive Excel report with formatting."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Summary Dashboard
    create_summary_sheet(wb, projects)
    
    # Create Detailed Projects Sheet
    create_projects_sheet(wb, projects)
    
    # Create Individual sheets for each project
    for i, project in enumerate(projects[:10]):  # Limit to first 10 for performance
        create_project_sheet(wb, project, i + 1)
    
    # Save the workbook
    wb.save(output_path)
    return output_path


def create_summary_sheet(wb, projects):
    """Create a summary dashboard sheet."""
    ws = wb.create_sheet(title="Dashboard")
    
    # Define styles
    header_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E07020", end_color="E07020", fill_type="solid")
    
    subheader_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    
    data_font = Font(name='Arial', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = 'PMO Project Status Dashboard'
    ws['A1'].font = Font(name='Arial', size=20, bold=True, color="E07020")
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Report Date
    ws['A2'] = f'Report Date: {datetime.now().strftime("%d/%m/%Y")}'
    ws['A2'].font = Font(name='Arial', size=12)
    ws.merge_cells('A2:G2')
    
    # Summary Statistics
    ws['A4'] = 'Summary Statistics'
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    ws.merge_cells('A4:C4')
    
    # Stats
    total_projects = len(projects)
    on_track = sum(1 for p in projects if 'on track' in str(p.get('health', '')).lower())
    at_risk = sum(1 for p in projects if 'at risk' in str(p.get('health', '')).lower())
    off_track = sum(1 for p in projects if 'off track' in str(p.get('health', '')).lower() or 'delayed' in str(p.get('health', '')).lower())
    
    stats_data = [
        ['Metric', 'Value'],
        ['Total Projects', total_projects],
        ['On Track', on_track],
        ['At Risk', at_risk],
        ['Off Track/Delayed', off_track],
        ['', ''],
        ['Total Budget', f"{sum(p.get('budget_total', 0) for p in projects):,.2f} SAR"],
        ['Total Spent', f"{sum(p.get('budget_spent', 0) for p in projects):,.2f} SAR"],
        ['Average Progress', f"{sum(p.get('timeline_actual', 0) for p in projects) / len(projects):.1f}%"]
    ]
    
    for row_num, row_data in enumerate(stats_data, 5):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 5:  # Header row
                cell.font = subheader_font
                cell.fill = subheader_fill
            else:
                cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Projects by Health Status
    ws['E4'] = 'Health Status Overview'
    ws['E4'].font = header_font
    ws['E4'].fill = header_fill
    ws.merge_cells('E4:G4')
    
    health_data = [
        ['Status', 'Count', 'Percentage'],
        ['On Track', on_track, f"{(on_track/total_projects*100):.1f}%"],
        ['At Risk', at_risk, f"{(at_risk/total_projects*100):.1f}%"],
        ['Off Track', off_track, f"{(off_track/total_projects*100):.1f}%"]
    ]
    
    for row_num, row_data in enumerate(health_data, 5):
        for col_num, value in enumerate(row_data, 5):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 5:  # Header row
                cell.font = subheader_font
                cell.fill = subheader_fill
            else:
                cell.font = data_font
                # Color code based on status
                if col_num == 5:  # Status column
                    if value == 'On Track':
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value == 'At Risk':
                        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                    elif value == 'Off Track':
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20


def create_projects_sheet(wb, projects):
    """Create a detailed sheet with all projects."""
    ws = wb.create_sheet(title="All Projects")
    
    # Headers
    headers = [
        'Project #',
        'Project Name',
        'Category',
        'Status',
        'Health',
        'GM',
        'Director',
        'Lead',
        'End Date',
        'Days Remaining',
        'Total Budget (SAR)',
        'Spent (SAR)',
        'Spent %',
        'Remaining (SAR)',
        'Timeline Actual %',
        'Timeline Planned %',
        'Schedule Variance %',
        'Current Activities',
        'Risks',
        'Issues'
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_num, project in enumerate(projects, 2):
        data_row = [
            project.get('number', ''),
            project.get('name', ''),
            project.get('category', ''),
            project.get('status', ''),
            project.get('health', ''),
            project.get('gm', ''),
            project.get('director', ''),
            project.get('operational_lead', ''),
            project.get('contract_end_date', ''),
            project.get('days_remaining', 0),
            project.get('budget_total', 0),
            project.get('budget_spent', 0),
            project.get('budget_spent_pct', 0),
            project.get('budget_remaining', 0),
            project.get('timeline_actual', 0),
            project.get('timeline_planned', 0),
            project.get('schedule_variance', 0),
            project.get('current_activities', '')[:200],  # Truncate long text
            project.get('risks', '')[:200],
            project.get('issues', '')[:200]
        ]
        
        for col_num, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            
            # Format numbers
            if col_num in [11, 12, 14]:  # Budget columns
                cell.number_format = '#,##0.00'
            elif col_num in [13, 15, 16, 17]:  # Percentage columns
                if isinstance(value, (int, float)):
                    cell.value = value / 100 if col_num != 17 else value / 100
                    cell.number_format = '0.0%'
            
            # Color code health status
            if col_num == 5:  # Health column
                health = str(value).lower()
                if 'on track' in health:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                elif 'at risk' in health:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C5700")
                elif 'off track' in health or 'delayed' in health:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
            
            # Color code variance
            if col_num == 17:  # Variance column
                if isinstance(value, (int, float)):
                    if value > 0:
                        cell.font = Font(color="006100")
                    elif value < -5:
                        cell.font = Font(color="9C0006")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add autofilter
    ws.auto_filter.ref = ws.dimensions
    
    # Freeze top row
    ws.freeze_panes = 'A2'


def create_project_sheet(wb, project, index):
    """Create individual sheet for a project."""
    # Sanitize sheet name (Excel has 31 char limit)
    sheet_name = f"{index}. {project['name'][:25]}"
    ws = wb.create_sheet(title=sheet_name)
    
    # Title
    ws['A1'] = f"Project Report: {project['name']}"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color="E07020")
    ws.merge_cells('A1:D1')
    
    # Basic Information
    info_data = [
        ['Field', 'Value'],
        ['Project Name', project.get('name', '')],
        ['Category', project.get('category', '')],
        ['Status', project.get('status', '')],
        ['Health', project.get('health', '')],
        ['Sponsor GM', project.get('gm', '')],
        ['Director', project.get('director', '')],
        ['Project Lead', project.get('operational_lead', '')],
        ['Vendor', project.get('vendor', '')],
        ['Contract End Date', project.get('contract_end_date', '')],
        ['Days Remaining', project.get('days_remaining', 0)],
        ['', ''],
        ['Budget Metrics', ''],
        ['Total Budget', f"{project.get('budget_total', 0):,.2f} SAR"],
        ['Spent', f"{project.get('budget_spent', 0):,.2f} SAR ({project.get('budget_spent_pct', 0):.1f}%)"],
        ['Remaining', f"{project.get('budget_remaining', 0):,.2f} SAR ({project.get('budget_remaining_pct', 0):.1f}%)"],
        ['', ''],
        ['Timeline Metrics', ''],
        ['Actual Progress', f"{project.get('timeline_actual', 0):.1f}%"],
        ['Planned Progress', f"{project.get('timeline_planned', 0):.1f}%"],
        ['Schedule Variance', f"{project.get('schedule_variance', 0):+.1f}%"],
        ['', ''],
        ['Current Activities', project.get('current_activities', '[To be provided]')],
        ['Future Activities', project.get('future_activities', '[To be provided]')],
        ['Risks', project.get('risks', '[To be provided]')],
        ['Issues', project.get('issues', '[To be provided]')],
        ['Comments', project.get('comments', '[To be provided]')]
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    label_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write data
    for row_num, row_data in enumerate(info_data, 3):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            
            # Apply styles
            if row_num == 3:  # Header row
                cell.font = header_font
                cell.fill = header_fill
            elif col_num == 1:  # Label column
                cell.font = label_font
            
            # Color health status
            if row_data[0] == 'Health':
                health = str(value).lower()
                if 'on track' in health:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif 'at risk' in health:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif 'off track' in health or 'delayed' in health:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Color variance
            if row_data[0] == 'Schedule Variance':
                try:
                    variance_val = float(str(value).replace('%', '').replace('+', ''))
                    if variance_val > 0:
                        cell.font = Font(color="006100", bold=True)
                    elif variance_val < -5:
                        cell.font = Font(color="9C0006", bold=True)
                except:
                    pass
            
            cell.border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    
    # Merge cells for long text
    for row in range(25, 30):
        ws.merge_cells(f'B{row}:D{row}')


def generate_individual_excel_report(project, output_path):
    """Generate an Excel report for a single project."""
    wb = Workbook()
    wb.remove(wb.active)
    create_project_sheet(wb, project, 1)
    wb.save(output_path)
    return output_path
